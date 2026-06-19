"""Import opening balances (Saldo_początkowe) from tbl_dostawcy_LIB.xlsx.

File columns: Dostawca | Jednostka | Obszar | Saldo_poczatkowe
Sets Pallet.opening_balance per (supplier, area, unit). Missing master data and
pallet rows are created; zero values are skipped (default opening is already 0).

Usage:
    python -m scripts.import_opening_balances [--file tbl_dostawcy_LIB.xlsx]
"""
import argparse
import asyncio
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import insert, select

from app.core.database import AsyncSessionLocal
from app.models.area import Area
from app.models.pallet import Pallet
from app.models.supplier import Supplier
from app.models.unit import Unit


def parse_opening(path: str) -> list[tuple]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(v is None for v in r):
            continue
        supplier = str(r[0]).strip() if r[0] is not None else ""
        unit = str(r[1]).strip() if r[1] is not None else ""
        area = str(r[2]).strip() if r[2] is not None else ""
        try:
            opening = int(r[3])
        except (TypeError, ValueError):
            opening = None
        rows.append((supplier, unit, area, opening))

    wb.close()
    return rows


async def _name_map(session, model):
    return {m.name: m.uuid for m in (await session.execute(select(model))).scalars()}


async def run_import_opening(session, rows) -> dict:
    suppliers = await _name_map(session, Supplier)
    units = await _name_map(session, Unit)
    areas = await _name_map(session, Area)

    # Create any master data referenced by the file but not yet in the DB.
    created = {"suppliers": 0, "units": 0, "areas": 0}
    for name in {r[0] for r in rows if r[0]} - set(suppliers):
        session.add(Supplier(name=name))
        created["suppliers"] += 1
    for name in {r[1] for r in rows if r[1]} - set(units):
        session.add(Unit(name=name))
        created["units"] += 1
    for name in {r[2] for r in rows if r[2]} - set(areas):
        session.add(Area(name=name))
        created["areas"] += 1
    await session.flush()

    suppliers = await _name_map(session, Supplier)
    units = await _name_map(session, Unit)
    areas = await _name_map(session, Area)

    pallets = {
        (p.supplier_uuid, p.area_uuid, p.unit_uuid): p
        for p in (await session.execute(select(Pallet))).scalars()
    }

    applied = 0
    skipped_zero = 0
    skipped_bad = 0
    new_pallets = []

    for supplier, unit, area, opening in rows:
        if opening is None or not (supplier and unit and area):
            skipped_bad += 1
            continue
        if opening == 0:
            skipped_zero += 1
            continue

        key = (suppliers[supplier], areas[area], units[unit])
        pallet = pallets.get(key)
        if pallet is not None:
            pallet.opening_balance = opening
        else:
            new_pallets.append({
                "uuid": uuid4(),
                "supplier_uuid": key[0],
                "area_uuid": key[1],
                "unit_uuid": key[2],
                "quantity": 0,
                "opening_balance": opening,
            })
            pallets[key] = None  # reserve the key so duplicates update, not duplicate
        applied += 1

    if new_pallets:
        await session.execute(insert(Pallet), new_pallets)
    await session.commit()

    return {
        "applied": applied,
        "skipped_zero": skipped_zero,
        "skipped_bad": skipped_bad,
        "created_masterdata": created,
    }


async def _main(path: str) -> None:
    rows = parse_opening(path)
    async with AsyncSessionLocal() as session:
        result = await run_import_opening(session, rows)

    c = result["created_masterdata"]
    print("=" * 50)
    print("IMPORT SALDA POCZĄTKOWEGO — PODSUMOWANIE")
    print("=" * 50)
    print(f"Ustawiono saldo początkowe:   {result['applied']}")
    print(f"Pominięto (zero):             {result['skipped_zero']}")
    print(f"Pominięto (błędne wiersze):   {result['skipped_bad']}")
    print(f"Utworzono masterdata:         suppliers={c['suppliers']}, units={c['units']}, areas={c['areas']}")
    print("=" * 50)


def main() -> None:
    parser = argparse.ArgumentParser(description="Import opening balances from the suppliers library file.")
    parser.add_argument("--file", default="tbl_dostawcy_LIB.xlsx")
    args = parser.parse_args()
    asyncio.run(_main(args.file))


if __name__ == "__main__":
    main()
