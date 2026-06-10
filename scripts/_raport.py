"""Shared parsing/validation for the tbl_raport.xlsx import.

Column layout:
    Data_dodania | Data | Pracownik | Dostawca | Operacja | Wartosc | Jednostka | Obszar | Komentarz
"""
from dataclasses import dataclass
from datetime import date, datetime

from openpyxl import load_workbook

# Source operation -> our TransactionType value.
OP_MAP = {"IN": "RECEIPT", "OUT": "ISSUE", "KOREKTA": "CORRECTION"}


@dataclass
class Row:
    row_no: int
    created_at: datetime | None      # Data_dodania (auto timestamp in source)
    operation_date: date | None      # Data (user-selected)
    pracownik: str
    dostawca: str
    operacja: str
    wartosc: object
    jednostka: str
    obszar: str
    komentarz: str | None


def _s(value) -> str:
    return str(value).strip() if value is not None else ""


def parse_workbook(path: str) -> list[Row]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows: list[Row] = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if r is None or all(v is None or _s(v) == "" for v in r):
            continue

        created_at = r[0] if isinstance(r[0], datetime) else None
        raw_date = r[1]
        if isinstance(raw_date, datetime):
            op_date = raw_date.date()
        elif isinstance(raw_date, date):
            op_date = raw_date
        else:
            op_date = None

        rows.append(
            Row(
                row_no=i,
                created_at=created_at,
                operation_date=op_date,
                pracownik=_s(r[2]),
                dostawca=_s(r[3]),
                operacja=_s(r[4]).upper(),
                wartosc=r[5],
                jednostka=_s(r[6]),
                obszar=_s(r[7]),
                komentarz=_s(r[8]) or None,
            )
        )

    wb.close()
    return rows


def row_errors(row: Row) -> list[str]:
    """Return a list of reasons the row cannot become a transaction (empty = OK)."""
    errors = []
    if not row.pracownik:
        errors.append("brak Pracownik")
    if not row.dostawca:
        errors.append("brak Dostawca")
    if not row.jednostka:
        errors.append("brak Jednostka")
    if not row.obszar:
        errors.append("brak Obszar")
    if row.operacja not in OP_MAP:
        errors.append(f"nieznana Operacja '{row.operacja}'")
    try:
        int(row.wartosc)
    except (TypeError, ValueError):
        errors.append(f"Wartosc nie jest liczbą: {row.wartosc!r}")
    if row.created_at is None:
        errors.append("brak/zła Data_dodania")
    if row.operation_date is None:
        errors.append("brak/zła Data")
    return errors


def quantity_of(row: Row) -> int:
    # Source uses signed values (IN is negative); we store a positive count.
    return abs(int(row.wartosc))
