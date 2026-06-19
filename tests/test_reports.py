# tests/test_reports.py
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.dependencies import get_current_user
from app.main import app
from app.repositories.area_repo import AreaRepository
from app.repositories.pallet_repo import PalletRepository
from app.repositories.supplier_repo import SupplierRepository
from app.repositories.unit_repo import UnitRepository
from app.repositories.user_repo import UserRepository
from app.repositories.role_repo import RoleRepository
from app.services.transaction_service import TransactionService

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def make_user(role: str):
    return type(
        "U", (), {
            "uuid": uuid4(), "username": "u", "is_active": True,
            "role": type("R", (), {"uuid": uuid4(), "name": role})(),
        },
    )()


async def _seed(session):
    supplier = await SupplierRepository(session).create_one({"name": f"Sup {uuid4().hex[:6]}"})
    area = await AreaRepository(session).create_one({"name": f"Area {uuid4().hex[:6]}"})
    unit = await UnitRepository(session).create_one({"name": f"U {uuid4().hex[:6]}"})
    role = await RoleRepository(session).create_one({"name": f"r{uuid4().hex[:6]}"})
    user = await UserRepository(session).create_one(
        {"username": f"emp{uuid4().hex[:6]}", "hashed_password": "x", "role_uuid": role.uuid}
    )

    svc = TransactionService(session)
    base = {"supplier_uuid": supplier.uuid, "area_uuid": area.uuid, "unit_uuid": unit.uuid}
    await svc.create_transaction({**base, "type": "RECEIPT", "quantity": 10, "comment": None}, user.uuid)
    await svc.create_transaction({**base, "type": "ISSUE", "quantity": 3, "comment": None}, user.uuid)
    await session.commit()
    return supplier, area, unit


class TestReportsAPI:
    @pytest.mark.asyncio
    async def test_saldo_xlsx(self, client, session):
        await _seed(session)

        res = await client.get("/reports/saldo")
        assert res.status_code == 200
        assert res.headers["content-type"] == XLSX_MIME

        ws = load_workbook(BytesIO(res.content)).active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("Dostawca", "Jednostka", "Obszar", "Saldo_początkowe",
                           "IN", "OUT", "Korekty", "Saldo_końcowe", "Saldo-2%", "Saldo-1%")
        data = rows[1]
        assert data[3] == 0   # Saldo_początkowe (no opening set)
        assert data[4] == 10  # IN
        assert data[5] == 3   # OUT
        assert data[7] == -7  # Saldo_końcowe: opening 0 + (-10 + 3)

    @pytest.mark.asyncio
    async def test_saldo_with_opening_balance(self, client, session):
        supplier, area, unit = await _seed(session)  # movements → quantity -7
        pallet = await PalletRepository(session).get_stock(supplier.uuid, area.uuid, unit.uuid)
        pallet.opening_balance = -100
        await session.commit()

        ws = load_workbook(BytesIO((await client.get("/reports/saldo")).content)).active
        data = list(ws.iter_rows(values_only=True))[1]
        assert data[3] == -100   # Saldo_początkowe
        assert data[7] == -107   # Saldo_końcowe = -100 + (-7)

    @pytest.mark.asyncio
    async def test_ksiegowania_xlsx(self, client, session):
        await _seed(session)

        res = await client.get("/reports/ksiegowania")
        assert res.status_code == 200
        assert res.headers["content-type"] == XLSX_MIME

        ws = load_workbook(BytesIO(res.content)).active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "Data_dodania"
        # two transactions seeded
        assert len(rows) == 3

    @pytest.mark.asyncio
    async def test_history_uses_operation_date(self, client, session):
        from datetime import date

        supplier = await SupplierRepository(session).create_one({"name": f"S {uuid4().hex[:6]}"})
        area = await AreaRepository(session).create_one({"name": f"A {uuid4().hex[:6]}"})
        unit = await UnitRepository(session).create_one({"name": f"U {uuid4().hex[:6]}"})
        role = await RoleRepository(session).create_one({"name": f"r{uuid4().hex[:6]}"})
        user = await UserRepository(session).create_one(
            {"username": f"e{uuid4().hex[:6]}", "hashed_password": "x", "role_uuid": role.uuid}
        )

        await TransactionService(session).create_transaction(
            {
                "type": "RECEIPT", "supplier_uuid": supplier.uuid, "area_uuid": area.uuid,
                "unit_uuid": unit.uuid, "quantity": 5, "comment": None, "date": date(2026, 1, 15),
            },
            user.uuid,
        )
        await session.commit()

        ws = load_workbook(BytesIO((await client.get("/reports/ksiegowania")).content)).active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "Data_dodania" and rows[0][1] == "Data"
        assert str(rows[1][1])[:10] == "2026-01-15"  # Data = selected operation date

        # Date range filters by operation_date (Jan row excluded from a Feb window).
        ranged = await client.get("/reports/ksiegowania?start=2026-02-01&end=2026-02-28")
        ws2 = load_workbook(BytesIO(ranged.content)).active
        assert len(list(ws2.iter_rows(values_only=True))) == 1  # header only

    @pytest.mark.asyncio
    async def test_reports_require_permission(self, client):
        app.dependency_overrides[get_current_user] = lambda: make_user("nobody")
        res = await client.get("/reports/saldo")
        assert res.status_code == 403
